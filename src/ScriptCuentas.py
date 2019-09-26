import argparse
import ast
from datetime import datetime
from functools import reduce
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


__author__ = 'bernatfelip'


# ----------------- PARSING/FORMATTING METHODS ----------------- #

def formatDate(date):
    return datetime.fromtimestamp(date/1000).strftime('%d/%m/%Y')


def formatAmount(amount):
    if(amount == 0):
        return "0,00"
    amountStr = str(amount)
    return amountStr[:-2] + ',' + amountStr[-2:] + ' €'


def parseDate(date):
    """Parses a date in string
    format to a timestampt in millis

    Arguments:
        date {str} -- date in "DD/MMYY" format

    Returns:
        int -- Milliseconds timestamp
    """
    dt_obj = datetime.strptime(date, '%d/%m/%Y')
    return dt_obj.timestamp()*1000


def parseAmount(amount):
    """Parses amount in string
    format to cents of euro as an int.

    Arguments:
        amount {str} -- Euros with a comma delimiter,
                           two decimal points and an optional
                           minus sign

    Returns:
        int -- Cents of euro (negative if payment)
    """
    if(amount[0] == '-'):
        return -1*int(amount[1:].replace(',', '').replace('.',''))
    else:
        return int(amount.replace(',', '').replace('.',''))


def parseAdvance(advance):
    """Parses advance input in string to
    boolean

    Arguments:
        advance {str} -- 'Y' or 'N' string

    Returns:
        bool -- Returns true if 'Y', else 'N
    """
    return (advance == 'Y')


def parseTransactions(inputFileName):
    """Inputs all the data from the input file

    Arguments:
        inputFileName {str} -- Path of the file containing the transactions

    Returns:
        {
            movement: str,
            date: int,
            info: str,
            amount: int,
            name: str,
            event: str,
            concept: str,
            advance: bool,
            origin: str,
            comment: str
        }[] -- List of transactions fully parsed
        str[] -- List of distinct event names present
    """
    transactions = []
    eventNames = []
    with open(inputFileName, 'r') as inputFile:
        next(inputFile)  # 1st line should be avoided
        next(inputFile)  # 2nd line should be avoided also
        for line in inputFile:
            line[:-1]  # Remove \n character
            dataArray = line[:-1].split(';')
            transaction = {
                'movement': dataArray[0],
                'date': parseDate(dataArray[1]),
                'info': dataArray[2],
                'amount': parseAmount(dataArray[3]),
                'name': dataArray[4],
                'event': dataArray[5],
                'concept': dataArray[6],
                'advance': parseAdvance(dataArray[7]),
                'origin': dataArray[8],
                'comment': dataArray[9]
            }
            transactions.append(transaction)
            if(transaction['event'] not in eventNames):
                eventNames.append(transaction['event'])
    return transactions, eventNames


# --------------- CALCULATING METHODS --------------- #
def sumAmounts(transactions):
    """Sums the amounts of all the transactions

    Arguments:
        transactions -- transactions to sum amounts

    Returns:
        int -- sum of all the transactions amounts
    """
    if(len(transactions) > 0):
        return reduce(lambda x, y: x+y,
                      [transaction['amount'] for transaction in transactions])
    else:
        return 0


def calulcateGross(transactionsEventLoss, transactionsEventProfit, origin):
    """Calculates the loss, profit and balance of a certain
    origin (PAYPAL or CAIXA)

    Arguments:
        transactionsEventLoss {transaction list} -- Negative transactions
        transactionsEventProfit {transaction list} -- Positive transactions
        origin {string} -- [PAYPAL or CAIXA

    Returns:
        {int int int} -- loss profit and balance
    """
    transactionsLoss = list(filter(
        lambda transaction:
        transaction['origin'] == origin, transactionsEventLoss))
    transactionsProfit = list(filter(
        lambda transaction:
        transaction['origin'] == origin, transactionsEventProfit))

    grossLoss = sumAmounts(transactionsLoss)

    grossProfit = sumAmounts(transactionsProfit)

    grossBalance = grossLoss + grossProfit

    return {
        'grossLoss': grossLoss,
        'grossProfit': grossProfit,
        'grossBalance': grossBalance,
    }


def getEventConceptNames(transactions):
    """Returns concept names appearing in transaction list

    Arguments:
        transactions {transaction list} -- transactions to search

    Returns:
        string[] -- list of concept names
    """
    conceptNames = []
    for transaction in transactions:
        if(transaction['concept'] not in conceptNames):
            conceptNames.append(transaction['concept'])
    return conceptNames


def calculateEventLossConcepts(conceptNames, transactionsLoss):
    """Calculates for every concept the amount spent

    Arguments:
        conceptNames {string[]} -- Concept names presearched
        transactionsLoss {list of transactions} -- loss transactions to search

    Returns:
        calculated concept list -- each concept identified by name
                                   has its amount
    """
    concepts = []
    total = 0
    for conceptName in conceptNames:
        transactionsLossConcept = list(filter(
            lambda t: (t['concept'] == conceptName) and (not t['advance']),
            transactionsLoss))
        amount = sumAmounts(transactionsLossConcept)
        concepts.append({
            'name': conceptName,
            'amount': amount
        })
        total += amount
    return concepts, total


def myKey(transaction):
    return transaction['date']


def calculateEvent(eventName, transactions):
    """Does all the needed calulations for each event

    Arguments:
        eventName {string} -- name of the event
        transactions {transaction list} -- all the transaction

    Returns:
        {
        'transactionsProfit': list of transactions -- positive transactions
        'transactionsLoss': list of transactions -- negative transactions
        'caixaGross': object -- loss profit and balance in caixa
        'paypalGross': object -- loss profit and balance in paypal
        'totalGrossLoss': int -- total loss
        'totalGrossProfit':  int -- total profit
        'totalGrossBalance':  int -- total balance
        'netLossConcepts': array -- calculated concepts with each amount
        'netLossTotal': int -- total loss net
        }
    """
    transactionsEvent = list(filter(
        lambda transaction: transaction['event'] == eventName, transactions))
    transactionsEventLoss = list(filter(
        lambda transaction:
            transaction['amount'] < 0, transactionsEvent))
    transactionsEventProfit = list(filter(
        lambda transaction:
            transaction['amount'] > 0, transactionsEvent))
    caixaGross = calulcateGross(
        transactionsEventLoss, transactionsEventProfit, 'CAIXA')
    paypalGross = calulcateGross(
        transactionsEventLoss, transactionsEventProfit, 'PAYPAL')

    grossLoss = caixaGross['grossLoss'] + paypalGross['grossLoss']
    grossProfit = caixaGross['grossProfit'] + paypalGross['grossProfit']
    grossBalance = grossProfit + grossLoss

    conceptNames = getEventConceptNames(transactionsEventLoss)
    netLossConcepts, netLossTotal = calculateEventLossConcepts(
        conceptNames, transactionsEventLoss)

    return {
        'transactionsProfit': transactionsEventProfit,
        'transactionsLoss': transactionsEventLoss,
        'caixaGross': caixaGross,
        'paypalGross': paypalGross,
        'totalGrossLoss': grossLoss,
        'totalGrossProfit': grossProfit,
        'totalGrossBalance': grossBalance,
        'netLossConcepts': netLossConcepts,
        'netLossTotal': netLossTotal
    }


# ----------------- OUTPUT METHODS ----------------- #
def applyStyleHeader(cellPos, sheet):
    single = Side(border_style="medium", color="000000")
    cell = sheet[cellPos]
    cell.font = Font(name='Calibri', bold=True)
    cell.fill = PatternFill("solid", fgColor="9ff9ff")
    cell.border = Border(top=single, left=single, right=single, bottom=single)
    cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)


def applyStyleTransaction(cellPos, sheet):
    single = Side(border_style="thin", color="000000")
    cell = sheet[cellPos]
    cell.font = Font(name='Calibri', bold=False)
    cell.border = Border(top=single, left=single, right=single, bottom=single)
    cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)


def writeStructureBalance(sheet):
    sheet.merge_cells('A1:E1')
    sheet.merge_cells('F1:J1')
    sheet.merge_cells('M1:O1')
    sheet.merge_cells('L7:M7')
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 16
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['L'].width = 20
    sheet.column_dimensions['M'].width = 16
    sheet.column_dimensions['N'].width = 16
    sheet.column_dimensions['O'].width = 16
    sheet['L7'] = 'Net Loss'
    sheet['M1'] = 'Gross Balance'
    sheet['M2'] = 'Caixa'
    sheet['N2'] = 'Paypal'
    sheet['O2'] = 'Total'
    sheet['L3'] = 'Loss'
    sheet['L4'] = 'Profit'
    sheet['L5'] = 'Total'
    sheet['A1'] = 'Loss'
    sheet['A2'] = 'Name'
    sheet['B2'] = 'Concept'
    sheet['C2'] = 'Date'
    sheet['D2'] = 'Amount'
    sheet['E2'] = 'Comment'
    sheet['F1'] = 'Profit'
    sheet['F2'] = 'Name'
    sheet['G2'] = 'Concept'
    sheet['H2'] = 'Date'
    sheet['I2'] = 'Amount'
    sheet['J2'] = 'Comment'
    applyStyleHeader('A1', sheet)
    applyStyleHeader('A2', sheet)
    applyStyleHeader('B2', sheet)
    applyStyleHeader('C2', sheet)
    applyStyleHeader('D2', sheet)
    applyStyleHeader('E2', sheet)
    applyStyleHeader('F1', sheet)
    applyStyleHeader('F2', sheet)
    applyStyleHeader('G2', sheet)
    applyStyleHeader('H2', sheet)
    applyStyleHeader('I2', sheet)
    applyStyleHeader('J2', sheet)
    applyStyleHeader('M1', sheet)
    applyStyleHeader('M2', sheet)
    applyStyleHeader('N2', sheet)
    applyStyleHeader('O2', sheet)
    applyStyleHeader('L3', sheet)
    applyStyleHeader('L4', sheet)
    applyStyleHeader('L5', sheet)
    applyStyleHeader('L7', sheet)
    applyStyleHeader('M7', sheet)


def printTrans(transactions, col, sheet):
    transactions.sort(key=lambda t: t['date'])
    i = 3
    c1 = ''
    c2 = ''
    c3 = ''
    c4 = ''
    c5 = ''
    if(col == 'A'):
        c1 = 'A'
        c2 = 'B'
        c3 = 'C'
        c4 = 'D'
        c5 = 'E'
    else:
        c1 = 'F'
        c2 = 'G'
        c3 = 'H'
        c4 = 'I'
        c5 = 'J'

    for transaction in transactions:
        sheet[c1+str(i)] = transaction['name']
        sheet[c2+str(i)] = transaction['concept']
        sheet[c3+str(i)] = formatDate(transaction['date'])
        sheet[c4+str(i)] = formatAmount(transaction['amount'])
        sheet[c5+str(i)] = transaction['comment']
        applyStyleTransaction(c1+str(i), sheet)
        applyStyleTransaction(c2+str(i), sheet)
        applyStyleTransaction(c3+str(i), sheet)
        applyStyleTransaction(c4+str(i), sheet)
        applyStyleTransaction(c5+str(i), sheet)
        i += 1


def writeEvents(eventNames, balances, sheet, taxNames, grantNames):
    sheet['A1'] = 'Events'
    sheet['B1'] = 'Caixa'
    sheet['C1'] = 'Paypal'
    sheet['D1'] = 'Total'

    applyStyleHeader('A1', sheet)
    applyStyleHeader('B1', sheet)
    applyStyleHeader('C1', sheet)
    applyStyleHeader('D1', sheet)

    sheet.column_dimensions['A'].width = 24
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    i = 2
    totalCaixa = 0
    totalPaypal = 0
    total = 0
    for eventName in eventNames:
        if((eventName not in taxNames) and (eventName not in grantNames)):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['caixaGross']['grossBalance'])
            sheet['C'+str(i)] = formatAmount(balances[eventName]
                                             ['paypalGross']['grossBalance'])
            sheet['D'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossBalance'])
            totalCaixa += balances[eventName]['caixaGross']['grossBalance']
            totalPaypal += balances[eventName]['paypalGross']['grossBalance']
            total += balances[eventName]['totalGrossBalance']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            applyStyleTransaction('C'+str(i), sheet)
            applyStyleTransaction('D'+str(i), sheet)
            i += 1
    i += 2
    sheet['A'+str(i)] = 'Grants'
    applyStyleHeader('A'+str(i), sheet)
    i += 1
    for eventName in eventNames:
        if(eventName in grantNames):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['caixaGross']['grossBalance'])
            sheet['C'+str(i)] = formatAmount(balances[eventName]
                                             ['paypalGross']['grossBalance'])
            sheet['D'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossBalance'])
            totalCaixa += balances[eventName]['caixaGross']['grossBalance']
            totalPaypal += balances[eventName]['paypalGross']['grossBalance']
            total += balances[eventName]['totalGrossBalance']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            applyStyleTransaction('C'+str(i), sheet)
            applyStyleTransaction('D'+str(i), sheet)
            i += 1

    i += 2
    sheet['A'+str(i)] = 'Gross Total'
    applyStyleHeader('A'+str(i), sheet)
    sheet['B'+str(i)] = formatAmount(totalCaixa)
    sheet['C'+str(i)] = formatAmount(totalPaypal)
    sheet['D'+str(i)] = formatAmount(total)
    applyStyleTransaction('B'+str(i), sheet)
    applyStyleTransaction('C'+str(i), sheet)
    applyStyleTransaction('D'+str(i), sheet)
    i += 2
    sheet['A'+str(i)] = 'Taxes'
    applyStyleHeader('A'+str(i), sheet)
    i += 1
    for eventName in eventNames:
        if(eventName in taxNames):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['caixaGross']['grossBalance'])
            sheet['C'+str(i)] = formatAmount(balances[eventName]
                                             ['paypalGross']['grossBalance'])
            sheet['D'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossBalance'])
            totalCaixa += balances[eventName]['caixaGross']['grossBalance']
            totalPaypal += balances[eventName]['paypalGross']['grossBalance']
            total += balances[eventName]['totalGrossBalance']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            applyStyleTransaction('C'+str(i), sheet)
            applyStyleTransaction('D'+str(i), sheet)
            i += 1
    i += 2
    sheet['A'+str(i)] = 'Net Total'
    applyStyleHeader('A'+str(i), sheet)
    sheet['B'+str(i)] = formatAmount(totalCaixa)
    sheet['C'+str(i)] = formatAmount(totalPaypal)
    sheet['D'+str(i)] = formatAmount(total)
    applyStyleTransaction('B'+str(i), sheet)
    applyStyleTransaction('C'+str(i), sheet)
    applyStyleTransaction('D'+str(i), sheet)

    i += 4

    initialBank = int(input("Input initial bank balance in cents\n"))
    finalBank = int(input("Input final bank balance in cents\n"))
    initialPaypal = int(input("Input initial paypal balance in cents\n"))
    finalPaypal = int(input("Input final paypal balance in cents\n"))

    sheet['B'+str(i)] = 'Caixa'
    sheet['C'+str(i)] = 'Paypal'
    sheet['D'+str(i)] = 'Total'

    applyStyleHeader('B'+str(i), sheet)
    applyStyleHeader('C'+str(i), sheet)
    applyStyleHeader('D'+str(i), sheet)

    i += 1

    sheet['A'+str(i)] = 'Initial liquidity'
    sheet['B'+str(i)] = formatAmount(initialBank)
    sheet['C'+str(i)] = formatAmount(initialPaypal)
    sheet['D'+str(i)] = formatAmount(initialPaypal+initialBank)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)
    applyStyleTransaction('C'+str(i), sheet)
    applyStyleTransaction('D'+str(i), sheet)

    i += 1

    sheet['A'+str(i)] = 'Final liquidity'
    sheet['B'+str(i)] = formatAmount(finalBank)
    sheet['C'+str(i)] = formatAmount(finalPaypal)
    sheet['D'+str(i)] = formatAmount(finalPaypal+finalBank)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)
    applyStyleTransaction('C'+str(i), sheet)
    applyStyleTransaction('D'+str(i), sheet)

    i += 1

    sheet['A'+str(i)] = 'Theoretical difference'
    sheet['B'+str(i)] = formatAmount(finalBank-initialBank)
    sheet['C'+str(i)] = formatAmount(finalPaypal-initialPaypal)
    sheet['D'+str(i)] = formatAmount(finalPaypal +
                                     finalBank-initialBank-initialPaypal)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)
    applyStyleTransaction('C'+str(i), sheet)
    applyStyleTransaction('D'+str(i), sheet)

    i += 1

    sheet['A'+str(i)] = 'Real difference'
    sheet['B'+str(i)] = formatAmount(totalCaixa)
    sheet['C'+str(i)] = formatAmount(totalPaypal)
    sheet['D'+str(i)] = formatAmount(total)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)
    applyStyleTransaction('C'+str(i), sheet)
    applyStyleTransaction('D'+str(i), sheet)


def writeLossProfit(eventNames, balances, sheet, taxNames, grantNames):
    sheet['A1'] = 'Net Import [255]'
    sheet['B1'] = 'Sells [256]'

    applyStyleHeader('A1', sheet)
    applyStyleHeader('B1', sheet)

    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    i = 2
    totalProfit = 0
    for eventName in eventNames:
        if((eventName not in taxNames) and (eventName not in grantNames)):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossProfit'])
            totalProfit += balances[eventName]['totalGrossProfit']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            i += 1
    i += 1
    sheet['A'+str(i)] = 'Other incomes [265]'
    sheet['B'+str(i)] = 'Grants'
    applyStyleHeader('A'+str(i), sheet)
    applyStyleHeader('B'+str(i), sheet)
    i += 1
    for eventName in eventNames:
        if(eventName in grantNames):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossProfit'])
            totalProfit += balances[eventName]['totalGrossProfit']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            i += 1
    i += 1
    sheet['A'+str(i)] = 'Income Total'
    sheet['B'+str(i)] = formatAmount(totalProfit)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)
    i += 3
    sheet['A'+str(i)] = 'Other expenses [279]'
    sheet['B'+str(i)] = 'Expenses'
    applyStyleHeader('A'+str(i), sheet)
    applyStyleHeader('B'+str(i), sheet)
    i += 1
    totalLoss = 0
    for eventName in eventNames:
        if((eventName not in taxNames) and (eventName not in grantNames)):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossLoss'])
            totalLoss += balances[eventName]['totalGrossLoss']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            i += 1
    i += 1
    sheet['A'+str(i)] = 'Loss Total'
    sheet['B'+str(i)] = formatAmount(totalLoss)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)
    i += 2
    sheet['A'+str(i)] = 'Result before tax [325]'
    applyStyleHeader('A'+str(i), sheet)
    sheet['B'+str(i)] = formatAmount(totalProfit+totalLoss)
    applyStyleTransaction('B'+str(i), sheet)
    i += 2
    sheet['A'+str(i)] = 'Taxes'
    sheet['B'+str(i)] = 'Expenses'
    applyStyleHeader('A'+str(i), sheet)
    applyStyleHeader('B'+str(i), sheet)
    i += 1
    for eventName in eventNames:
        if(eventName in taxNames):
            sheet['A'+str(i)] = eventName
            sheet['B'+str(i)] = formatAmount(balances[eventName]
                                             ['totalGrossBalance'])
            totalLoss += balances[eventName]['totalGrossBalance']
            applyStyleTransaction('A'+str(i), sheet)
            applyStyleTransaction('B'+str(i), sheet)
            i += 1
    i += 2
    sheet['A'+str(i)] = 'Account Result'
    sheet['B'+str(i)] = formatAmount(totalProfit+totalLoss)
    applyStyleHeader('A'+str(i), sheet)
    applyStyleTransaction('B'+str(i), sheet)


if __name__ == '__main__':
    """This script parses a csv file with the transactions
    and outputs the finances of the fiscal year.
    """

    # ----------------- PARSE DATA ----------------- #
    parser = argparse.ArgumentParser()
    parser.add_argument('--file', required=True,
                        help='Original filtered file to process (.csv)')
    parser.add_argument('--grants', required=True,
                        help='List of event names for grants')
    parser.add_argument('--taxes', required=True,
                        help='List of event names for taxes')
    args = parser.parse_args()
    inputFileName = str(args.file)
    grantNames = ast.literal_eval(args.grants)
    taxNames = ast.literal_eval(args.taxes)

    transactions, eventNames = parseTransactions(inputFileName)
    wb = Workbook()

    # ------------- CALCULATE BALANCES ------------- #
    balances = {}
    for eventName in eventNames:
        balances[eventName] = calculateEvent(eventName, transactions)
        sheet = wb.create_sheet(eventName)
        writeStructureBalance(sheet)
        printTrans(balances[eventName]['transactionsLoss'], 'A', sheet)
        printTrans(balances[eventName]['transactionsProfit'], 'F', sheet)
        sheet['M3'] = formatAmount(
            balances[eventName]['caixaGross']['grossLoss'])
        sheet['M4'] = formatAmount(
            balances[eventName]['caixaGross']['grossProfit'])
        sheet['M5'] = formatAmount(
            balances[eventName]['caixaGross']['grossBalance'])
        sheet['N3'] = formatAmount(
            balances[eventName]['paypalGross']['grossLoss'])
        sheet['N4'] = formatAmount(
            balances[eventName]['paypalGross']['grossProfit'])
        sheet['N5'] = formatAmount(
            balances[eventName]['paypalGross']['grossBalance'])
        sheet['O3'] = formatAmount(balances[eventName]['totalGrossLoss'])
        sheet['O4'] = formatAmount(balances[eventName]['totalGrossProfit'])
        sheet['O5'] = formatAmount(balances[eventName]['totalGrossBalance'])
        applyStyleTransaction('M3', sheet)
        applyStyleTransaction('M4', sheet)
        applyStyleTransaction('M5', sheet)
        applyStyleTransaction('N3', sheet)
        applyStyleTransaction('N4', sheet)
        applyStyleTransaction('N5', sheet)
        applyStyleTransaction('O3', sheet)
        applyStyleTransaction('O4', sheet)
        applyStyleTransaction('O5', sheet)
        i = 8
        for netLossConcept in balances[eventName]['netLossConcepts']:
            if(netLossConcept['amount'] != 0):
                sheet['L'+str(i)] = netLossConcept['name']
                sheet['M'+str(i)] = formatAmount(netLossConcept['amount'])
                applyStyleTransaction('M'+str(i), sheet)
                applyStyleTransaction('L'+str(i), sheet)
                i += 1
        sheet['L'+str(i)] = 'Total'
        sheet['M'+str(i)] = formatAmount(balances[eventName]['netLossTotal'])
        applyStyleHeader('L'+str(i), sheet)
        applyStyleTransaction('M'+str(i), sheet)

    # ------------- CALCULATE FINAL SUMMARY ------------- #
    sheet = wb.create_sheet("Final Summary")
    writeEvents(eventNames, balances, sheet, taxNames, grantNames)

    # ------------- CALCULATE LOSS & PROFIT ------------- #
    sheet = wb.create_sheet("Loss&Profit")
    writeLossProfit(eventNames, balances, sheet, taxNames, grantNames)

    wb.save("Cuentas.xls")
